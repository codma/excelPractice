from openpyxl import load_workbook
from openpyxl import Workbook

# 엑셀파일 쓰기
# write_wb = Workbook()
# 이름이 있는 시트를 생성
# write_ws = write_wb.create_sheet('생성시트')
# write_wb.save("/Users/jinjoolee/Desktop/Project/python/sample.xlsx")

# 입력 자리를 지정
# write_ws.cell(1,1,'연습해보자')

# 파일 불러오기
load_wb = load_workbook("/Users/jinjoolee/Desktop/Project/python/sample.xlsx")
load_ws = load_wb['생성시트']

# 셀 주소로 값 출력
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)


rowCount = 2
columnCount = 1

for row in all_values:
    for value in row:
        load_ws.cell(rowCount, columnCount, value)
        if rowCount == 2 and columnCount == 1:
            load_ws.cell(rowCount, columnCount, '깜짝등장')
        columnCount = columnCount + 1
    rowCount = rowCount + 1
    columnCount = 1


load_wb.save("/Users/jinjoolee/Desktop/Project/python/sample.xlsx")

